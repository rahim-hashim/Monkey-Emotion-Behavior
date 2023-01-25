import os
import numpy as np
from textwrap import indent
from pprint import pprint, pformat
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from markdown_print import weekday_dict

def read_header(ws):
	"""
	reads header row and returns dictionary of column names and their index

	Args:
		ws: openpyxl worksheet object
			
	Returns:
		headers: dictionary of column names and their index
	"""
	headers = {}
	header_row = list(ws.rows)[0]
	for c_index, cell in enumerate(header_row):
		headers[cell.value] = c_index
	return headers

def calculate_date(date):
	"""
	reads header row and returns dictionary of column names and their index

	Args:
		date: datetime object
			
	Returns:
		date_str: string of date in format 'January 01, 2022'
		weekday_str: string of weekday in format 'Monday'
	"""
	date_formatted = '20' + date
	datem = datetime.strptime(date, '%y%m%d')
	weekday = datem.weekday()
	weekday_str = weekday_dict[weekday]
	date_str = datem.strftime('%B %d, %Y')
	return date_str, weekday_str

def clean_rows(ws, headers):
	"""
	cleans empty rows of excel sheet

	Args:
		ws: openpyxl worksheet object
		headers: dictionary of column names and their index
			
	Returns:
		dates: list of dates in format 'yyyymmdd'
		sessions: list of session numbers
	"""
	dates = []; sessions = []
	empty_rows = []
	for r_index, row_cells in enumerate(ws.iter_rows(min_row=2)):
		if row_cells[headers['Date']].value != None:
			date = row_cells[headers['Date']].value
			task = row_cells[headers['Task']].value
			session = row_cells[headers['Session Number']].value
			dates.append(date)
			sessions.append('_'.join([date, task, str(session)]))
		else:
			empty_rows.append(r_index+2) # index is 1-based, so not including header
	for r_index in reversed(empty_rows):
		ws.delete_rows(r_index, amount=1)
	return dates, sessions

def behavior_summary(df, session_obj):
	TRIAL_THRESHOLD = 20
	df_correct = df[df['correct'] == 1]
	# only include trials after subject has seen fractal <TRIAL_THRESHOLD> number of times
	session_df_count = df_correct[df_correct['fractal_count_in_block'] > TRIAL_THRESHOLD]
	
	lick_mean = []
	blink_mean = []
	for valence in sorted(session_df_count['valence'].unique(), reverse=True):
		valence_df = session_df_count[session_df_count['valence'] == valence]
		lick_mean.append(valence_df['lick_duration'].mean())
		blink_mean.append(valence_df['blink_duration_offscreen'].mean())

	return lick_mean, blink_mean

def outcome_params(session_obj):
	reward_list = []
	airpuff_list = []
	num_reward_mag = len(session_obj.reward_outcome_params['reward_drops'])
	for reward_mag in range(num_reward_mag):
		reward_list.append(session_obj.reward_outcome_params['reward_drops'][reward_mag])
		reward_list.append(session_obj.reward_outcome_params['reward_freq'][reward_mag])
		reward_list.append(session_obj.reward_outcome_params['reward_length'][reward_mag])
	num_airpuff_mag = len(session_obj.airpuff_outcome_params['airpuff_mag'])
	for airpuff_mag in range(num_reward_mag):
		airpuff_list.append(session_obj.airpuff_outcome_params['airpuff_pulses'][airpuff_mag])
		airpuff_list.append(session_obj.airpuff_outcome_params['airpuff_freq'][airpuff_mag])
	return reward_list, airpuff_list

def write_to_excel(df, session_obj, path_obj, verbose=False):
	print('Writing to excel...')
	EXCEL_PATH = path_obj.EXCEL_PATH
	excel_path_file = os.path.basename(os.path.normpath(EXCEL_PATH))
	wb = load_workbook(filename = EXCEL_PATH)
	ws = wb['Data']
	headers = read_header(ws)
	date = df['date'].unique()[0]
	session_num = int(session_obj.session_num) + 1
	dates, sessions = clean_rows(ws, headers)
	if verbose:
		print(' Dates in {}:'.format(excel_path_file))
		print(indent(pformat(dates), '  '))
	date_str, weekday_str = calculate_date(date)
	session_task_num = '_'.join([date_str, session_obj.task, str(session_num)])
	if session_task_num not in sessions:
		print('  Writing {} to {}'.format(date, excel_path_file))
		start_datetime = df['trial_datetime_start'].iloc[0]
		end_datetime = df['trial_datetime_end'].iloc[-1]
		session_time = session_obj.session_length
		total_attempts_min = session_obj.total_attempts_min
		perc_initiated = session_obj.prop_trials_initiated
		perc_correct = session_obj.prop_correct_CS_on
		reward_list, airpuff_list = outcome_params(session_obj)
		lick_valence_list, blink_valence_list = behavior_summary(df, session_obj)
		behavior_values = [date_str, 
								weekday_str,
								session_obj.monkey,
								session_obj.task,
								session_num,
								start_datetime.time(),
								end_datetime.time(),
								session_time,
								len(df),
								np.sum(df['correct']),
								total_attempts_min,
								perc_initiated,
								perc_correct]
		all_data = behavior_values + \
							reward_list + airpuff_list + \
							lick_valence_list + blink_valence_list
		ws.append(all_data)
		wb.save(EXCEL_PATH)


